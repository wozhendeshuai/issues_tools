# 处理无法获得的PR，然后记录到文件中，记录格式如下，时间|PRNumber，时间|PR编号
import os
import time


# 查看是否存在该文件路径，不存在则创建
def path_exists_or_create(file_path):
    if not os.path.exists(file_path):
        print(" 不存在该路径   ", file_path)
        os.makedirs(file_path)
    print(" 已有该路径   ", file_path)


def write_exception_file(pr_number, project_name, exception, filename):
    current_path = os.getcwd() + '\\exception_data\\'  # 获取当前路径
    path_exists_or_create(current_path)
    #  print(current_path)
    path = current_path + filename  # 在当前路径创建名为test的文本文件
    now_time = time.strftime('%Y-%m-%d %H:%M:%S ', time.localtime(time.time()))  # 获取当前时间
    context = project_name + ", " + pr_number.__str__() + ', ' + now_time + ', ' + exception + "\n"
    if os.path.exists(path):
        print(path + ' is already exist')
        print('context is :' + context)
    else:
        print("创建文件哦：" + path)
    file = open(path, 'a+')
    file.write(context)
    file.close()


# 写入数据库时，去操作一下
# write_file(100, 'jsonjson', "404")

# print("当前时间： ", time.strftime('%Y-%m-%d %H:%M:%S ', time.localtime(time.time())))
# 将数据写入到csv文件中
def write_csv(repo_name, filename, issues_text, labels):
    current_path = os.getcwd() + '\\data\\' + repo_name + "\\"  # 获取当前路径
    path_exists_or_create(current_path)
    path = current_path + filename  # 在当前路径创建名为test的文本文件
    if os.path.exists(path):
        print(path + ' is already exist')
        print('context is :' + issues_text)
    else:
        print("创建文件哦：" + path)
    # 打开文件
    f = open(path, 'a+', encoding='utf-8')
    # 将信息写入到文件中
    f.write(issues_text + "," + str(labels) + "\n")
    # 关闭文件
    f.close()


import xlsxwriter as xw

import openpyxl


# 将数据写入到excel文件中,excel表头为issues_text,labels
def write_excel(repo_name, filename, issues_text, labels):
    current_path = os.getcwd() + '\\data\\' + repo_name + "\\"  # 获取当前路径
    path_exists_or_create(current_path)
    path = current_path + filename  # 在当前路径创建名为test的文本文件
    if os.path.exists(path):
        print(path + ' is already exist')
        data = openpyxl.load_workbook(path)
        print(data.get_sheet_names())  # 输出所有工作页的名称
        # 取第一张表
        sheetnames = data.get_sheet_names()
        table = data.get_sheet_by_name(sheetnames[0])
        table = data.active
        print(table.title)  # 输出表名
        nrows = table.max_row  # 获得行数
        table.cell(nrows + 1, 1).value = issues_text
        table.cell(nrows + 1, 2).value =  str(labels)
        data.save(path)
    else:
        workbook = xw.Workbook(path)  # 创建工作簿
        worksheet1 = workbook.add_worksheet("sheet1")  # 创建子表
        worksheet1.activate()  # 激活表
        title = ['issues_text', 'labels']  # 设置表头
        worksheet1.write_row('A1', title)  # 从A1单元格开始写入表头
        print("创建文件哦：" + path)
        worksheet1.write_row('A2', [issues_text, str(labels)])  # 从A2单元格开始写入数据
        workbook.close()  # 关闭表
